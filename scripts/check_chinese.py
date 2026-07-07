import openpyxl
import re

wb = openpyxl.load_workbook('frontend/public/demo-sample.xlsx', read_only=True, data_only=True)

zh_pattern = re.compile(r'[\u4e00-\u9fff\u3400-\u4dbf]')
found = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        for col_idx, cell in enumerate(row, 1):
            if cell and isinstance(cell, str) and zh_pattern.search(cell):
                key = f'{sheet_name}!R{row_idx}C{col_idx}'
                if len(found) < 80:
                    found[key] = cell[:80]

print(f'Found {len(found)} cells with Chinese characters:')
for k, v in sorted(found.items()):
    print(f'  {k}: {repr(v)}')
