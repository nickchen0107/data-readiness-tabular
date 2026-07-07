"""Fix remaining Chinese content in demo-sample.xlsx"""
import openpyxl
import re
import random

random.seed(42)

wb = openpyxl.load_workbook('frontend/public/demo-sample.xlsx')
ws = wb['ALL']

# 1. Fix supplier names (col 14) - replace Chinese with English
SUPPLIER_MAP = {
    '供應商A': 'Supplier A',
    '供應商B\n代工': 'Supplier B\n(OEM)',
    '供應商C': 'Supplier C',
    '代工廠D': 'OEM Partner D',
    '供應商E': 'Supplier E',
    'OEM廠F': 'OEM Partner F',
    '供應商G\n組裝': 'Supplier G\n(Assembly)',
}

# 2. Headers (row 1) col 16-43 translation
HEADER_MAP = {
    'USD成本': 'USD Cost',
    'USD 成本\n 單價(未稅)': 'USD Cost\nUnit Price (excl. tax)',
    '台幣NTD': 'NTD Amount',
    '台幣NTD\n單價(含稅)': 'NTD Unit Price\n(Tax incl.)',
    '台幣NTD\n總價(含稅)': 'NTD Total\n(Tax incl.)',
    '進出口報關單號/ 客戶發票號碼': 'Import/Export Declaration No.\n/ Customer Invoice No.',
    '國內發票匯率': 'Domestic Invoice\nExchange Rate',
    '國內發票號碼': 'Domestic Invoice No.',
    '國內發票(含稅)': 'Domestic Invoice\n(Tax incl.)',
    '國外匯款通知書': 'Overseas Remittance\nAdvice',
    '出口報單匯率': 'Export Declaration\nExchange Rate',
    '應收金額(NT)': 'Receivable\nAmount (NT)',
    '收款日期': 'Payment\nReceived Date',
    '實收金額(NT)': 'Actual Received\nAmount (NT)',
    '應收帳款狀態': 'A/R Status',
    '訂單狀態': 'Order Status',
    '類別': 'Category',
    '區域': 'Region',
    '業績獎金計算月份': 'Commission\nCalc. Month',
    '供應商匯款帳號': 'Supplier\nBank Account',
    '付款日期': 'Payment Date',
    '付款狀態': 'Payment Status',
    '發票日期': 'Invoice Date',
    '發票號碼': 'Invoice No.',
    '匯率': 'Exchange Rate',
    '備註': 'Remark',
}

# 3. Data value translations
VALUE_MAP = {
    '已清帳': 'Settled',
    '未清帳': 'Unsettled',
    '買賣': 'Trade',
    '國外': 'Overseas',
    '國內': 'Domestic',
    '已出貨': 'Shipped',
    '未出貨': 'Pending',
    '已付款': 'Paid',
    '未付款': 'Unpaid',
    '部分付款': 'Partial',
    '已開立': 'Issued',
    '作廢': 'Voided',
    '年': '/',
    '月底前': ' end of month',
    '月': '/',
    '底前': ' before end',
}

# 4. Fake bank accounts
FAKE_BANKS = [
    'Chase Bank 0012-3456-7890',
    'HSBC USD Acct: 081-001623-881',
    'Citibank Intl 4455-6677-8899',
    'Wells Fargo 9988-7766-5544',
    'Standard Chartered 052-0281-033',
]

zh_pattern = re.compile(r'[\u4e00-\u9fff\u3400-\u4dbf]')

print("Fixing headers...")
for col_idx in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col_idx)
    if cell.value and isinstance(cell.value, str):
        val = str(cell.value)
        for zh, en in HEADER_MAP.items():
            if zh in val:
                cell.value = en
                break

print("Fixing data cells...")
for row_idx in range(2, ws.max_row + 1):
    # Col 14: Supplier names
    cell = ws.cell(row=row_idx, column=14)
    if cell.value and isinstance(cell.value, str):
        val = str(cell.value)
        for zh, en in SUPPLIER_MAP.items():
            if zh == val or zh in val:
                cell.value = en
                break
        # If still Chinese, generic replace
        if zh_pattern.search(str(cell.value)):
            cell.value = f'Supplier {random.choice("ABCDEFG")}'

    # Col 22: Bank account - replace with fake
    cell = ws.cell(row=row_idx, column=22)
    if cell.value and isinstance(cell.value, str) and zh_pattern.search(cell.value):
        cell.value = random.choice(FAKE_BANKS)

    # All other cells: translate known Chinese values
    for col_idx in range(1, ws.max_column + 1):
        if col_idx in (14, 22):
            continue  # already handled
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value and isinstance(cell.value, str) and zh_pattern.search(cell.value):
            val = cell.value
            for zh, en in VALUE_MAP.items():
                val = val.replace(zh, en)
            # If still has Chinese after replacements, clear it
            if zh_pattern.search(val):
                # Keep the cell but remove Chinese chars, keep numbers/English
                val = re.sub(r'[\u4e00-\u9fff\u3400-\u4dbf]+', '', val).strip()
            cell.value = val if val else None

# Also fix OA List sheet (if it exists)
if 'OA List' in wb.sheetnames:
    ws2 = wb['OA List']
    print("Fixing OA List sheet...")
    for row_idx in range(1, ws2.max_row + 1):
        for col_idx in range(1, ws2.max_column + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            if cell.value and isinstance(cell.value, str) and zh_pattern.search(cell.value):
                val = cell.value
                for zh, en in VALUE_MAP.items():
                    val = val.replace(zh, en)
                for zh, en in SUPPLIER_MAP.items():
                    val = val.replace(zh, en)
                if zh_pattern.search(val):
                    val = re.sub(r'[\u4e00-\u9fff\u3400-\u4dbf]+', '', val).strip()
                cell.value = val if val else None
else:
    print(f"Sheets: {wb.sheetnames} - no OA List found, skipping")

print("Saving...")
wb.save('frontend/public/demo-sample.xlsx')
print("Done!")

# Verify
wb2 = openpyxl.load_workbook('frontend/public/demo-sample.xlsx', read_only=True, data_only=True)
count = 0
for sheet_name in wb2.sheetnames:
    ws = wb2[sheet_name]
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and zh_pattern.search(cell):
                count += 1
                if count <= 5:
                    print(f"  Still Chinese: {cell[:60]}")
print(f"Remaining Chinese cells: {count}")
