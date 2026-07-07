"""
基於「合併_戎克訂單系統(新).xlsx」製作模糊化的 demo 檔案。
保留完整 6600 行結構，只模糊化個資相關欄位。
"""
import openpyxl
import random
import os

random.seed(42)

# 模糊化用的假資料池
FAKE_CLIENTS = [
    'Alpha Corp\n(ALPHA)', 'Beta Systems', 'Gamma Electronics\n(GAMMA)',
    'Delta Tech', 'Epsilon IoT\n(EPS)', 'Zeta Solutions',
    'Eta Industries', 'Theta Computing', 'Iota Devices\n(IOTA)',
    'Kappa Networks', 'Lambda Micro', 'Mu Semiconductor',
]
FAKE_SALES = ['Alice', 'Bob', 'Carol', 'David', 'Eve', 'Frank', 'Grace', 'Henry']
FAKE_SUPPLIERS = [
    '供應商A', '供應商B\n代工', '供應商C', '代工廠D',
    '供應商E', 'OEM廠F', '供應商G\n組裝',
]
FAKE_MODELS = [
    'IoT-GW-X100', 'Edge-SRV-S200', 'Panel-PC-P300',
    'IPC-M400', 'Adapter-19V-120W', 'Cable-EU-Type',
    'Sensor-Module-S700', 'Controller-K800', 'NISE-3600E',
    'AI-Accelerator-A900', 'Fanless-PC-F500', 'DIN-Rail-D600',
]

# 用 hash 做 deterministic 替換（同一個原始值永遠對應同一個假值）
client_map = {}
sales_map = {}
supplier_map = {}
model_map = {}

def get_fake(original, fake_pool, mapping):
    if not original or not str(original).strip():
        return original
    key = str(original).strip()[:20]
    if key not in mapping:
        mapping[key] = random.choice(fake_pool)
    return mapping[key]

def anonymize_po(value):
    if not value:
        return value
    return f'PO-{random.randint(2000000, 9999999)}'

def anonymize_pi(value):
    if not value:
        return value
    return f'PI-{random.randint(20190101, 20251231)}'

def anonymize_erp(value):
    if not value:
        return value
    return f'WO-{random.randint(10000, 99999)}'

print("Loading workbook...")
src = os.path.join(os.path.dirname(__file__), '..', '測試資料', '合併_戎克訂單系統(新).xlsx')
wb = openpyxl.load_workbook(src)

# Sheet ALL: columns to anonymize (0-indexed)
# 1=Client, 2=PO, 3=PI, 8=Sales, 13=Supplier, 14=ERP
ws = wb['ALL']
print(f"Processing ALL sheet: {ws.max_row} rows x {ws.max_column} cols")

for row_idx in range(2, ws.max_row + 1):
    # Client (col B = index 2)
    cell = ws.cell(row=row_idx, column=2)
    if cell.value:
        cell.value = get_fake(cell.value, FAKE_CLIENTS, client_map)
    
    # PO No (col C = index 3)
    cell = ws.cell(row=row_idx, column=3)
    if cell.value:
        cell.value = anonymize_po(cell.value)
    
    # PI No (col D = index 4)
    cell = ws.cell(row=row_idx, column=4)
    if cell.value:
        cell.value = anonymize_pi(cell.value)
    
    # Model No (col E = index 5)
    cell = ws.cell(row=row_idx, column=5)
    if cell.value:
        cell.value = get_fake(cell.value, FAKE_MODELS, model_map)
    
    # Sales (col I = index 9)
    cell = ws.cell(row=row_idx, column=9)
    if cell.value:
        cell.value = get_fake(cell.value, FAKE_SALES, sales_map)
    
    # Col 13 header rename (抵達戎克/客戶 → 收到/抵達) - handle in data too
    cell = ws.cell(row=row_idx, column=13)
    if cell.value and '戎克' in str(cell.value):
        cell.value = str(cell.value).replace('戎克', '倉庫')
    
    # Supplier (col N = index 14)
    cell = ws.cell(row=row_idx, column=14)
    if cell.value:
        cell.value = get_fake(cell.value, FAKE_SUPPLIERS, supplier_map)
    
    # ERP (col O = index 15)
    cell = ws.cell(row=row_idx, column=15)
    if cell.value:
        cell.value = anonymize_erp(cell.value)

# Fix header row for column 13
header_cell = ws.cell(row=1, column=13)
if header_cell.value:
    header_cell.value = 'Received/Arrived'

# Translate ALL headers to English
HEADER_EN = {
    ' ': 'Date',
    'Client \n客戶.': 'Client',
    'PO No.\n採購單號': 'PO No.',
    'PI No.\n確認訂單': 'PI No.',
    'Model No.\n產品名': 'Model No.',
    "Q'nty\n數量": "Q'nty",
    '單價\n (含稅)': 'Unit Price\n(Tax incl.)',
    'Total Price \n總價 (含稅)': 'Total Price\n(Tax incl.)',
    'Sales\n業務': 'Sales',
    'Shipping \n運輸方式': 'Shipping',
    'Delivery \n出貨日': 'Delivery Date',
    'Pickup Location\n出貨地': 'Pickup Location',
    '抵達戎克\n/客戶': 'Received/Arrived',
    '收到/抵達': 'Received/Arrived',
    'Supplier\n供應商': 'Supplier',
    '廠商訂單/工單\n(ERP單號)': 'Work Order\n(ERP No.)',
}

for col_idx in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col_idx)
    if cell.value and str(cell.value) in HEADER_EN:
        cell.value = HEADER_EN[str(cell.value)]
    elif cell.value:
        # Remove Chinese from bilingual headers (keep English part)
        val = str(cell.value)
        if '\n' in val:
            parts = val.split('\n')
            # Keep the part that looks English (has ASCII letters)
            en_parts = [p.strip() for p in parts if any(c.isascii() and c.isalpha() for c in p)]
            if en_parts:
                cell.value = en_parts[0]

# Translate Chinese values in data cells to English equivalents
ZH_TO_EN = {
    '小計': 'Subtotal',
    '合計': 'Total',
    '同': 'Same as',
    '海運': 'Sea Freight',
    '空運': 'Air Freight',
    '快遞': 'Express',
    '自取': 'Self Pickup',
    '倉庫': 'Warehouse',
    '已出貨': 'Shipped',
    '未出貨': 'Not Shipped',
    '已付款': 'Paid',
    '未付款': 'Unpaid',
    '備註': 'Remark',
}

print("Translating Chinese values to English...")
for row_idx in range(2, ws.max_row + 1):
    for col_idx in range(1, min(ws.max_column + 1, 44)):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value and isinstance(cell.value, str):
            val = cell.value
            for zh, en in ZH_TO_EN.items():
                if zh in val:
                    val = val.replace(zh, en)
            if val != cell.value:
                cell.value = val

# Sheet OA List: anonymize client and sales
ws2 = wb['OA List']
print(f"Processing OA List sheet: {ws2.max_row} rows")

# Translate OA List headers
OA_HEADERS = {
    '日期': 'Date',
    'Client\n客戶': 'Client',
    'Sales \n業務': 'Sales',
    'USD O/A Amount\n美金放帳額度': 'USD O/A Amount',
    'Balanced Date\n結算日期': 'Balanced Date',
    'Remark\n備註': 'Remark',
}
for col_idx in range(1, ws2.max_column + 1):
    cell = ws2.cell(row=2, column=col_idx)  # OA List header is row 2
    if cell.value and str(cell.value) in OA_HEADERS:
        cell.value = OA_HEADERS[str(cell.value)]

# Translate row 1 note
note_cell = ws2.cell(row=1, column=1)
if note_cell.value:
    note_cell.value = "※ O/A over US$3000 requires sales approval before shipment"

for row_idx in range(3, ws2.max_row + 1):
    cell = ws2.cell(row=row_idx, column=2)
    if cell.value:
        cell.value = get_fake(cell.value, FAKE_CLIENTS, client_map)
    cell = ws2.cell(row=row_idx, column=3)
    if cell.value:
        cell.value = get_fake(cell.value, FAKE_SALES, sales_map)
    # Translate any Chinese in remarks
    for col_idx in range(4, ws2.max_column + 1):
        cell = ws2.cell(row=row_idx, column=col_idx)
        if cell.value and isinstance(cell.value, str):
            val = cell.value
            for zh, en in ZH_TO_EN.items():
                if zh in val:
                    val = val.replace(zh, en)
            if val != cell.value:
                cell.value = val

output = os.path.join(os.path.dirname(__file__), '..', 'frontend', 'public', 'demo-sample.xlsx')
print(f"Saving to {output}...")
wb.save(output)
size = os.path.getsize(output)
print(f"Done! File size: {size / 1024:.0f} KB ({size / 1024 / 1024:.1f} MB)")
